import win32com.client as win32
from PyPDF2 import PdfMerger
import os
import pythoncom  # Import this to help manage COM threading
from docx import Document
import sys
import openpyxl

# Example usage
pdf_output_folder = r"C:\Users\Lenovo\OneDrive\Projekty\Skamba\sablony\PBR"
merged_pdf_name = "outputek.pdf"

# Load the Word document
doc_path = r"C:\Users\Lenovo\OneDrive\Projekty\Skamba\sablony\PBR\D131_PBŘ_nevýrobní_TZ.docx"
excel_path = r"C:\Users\Lenovo\OneDrive\Projekty\Skamba\sablony\PBR\D131_PBŘ_XXX_přílohy.xlsx"
doc = Document(doc_path)

# Define the target heading and level
target_heading = "Přílohy"
in_target_section = False
target_level = None  # We'll capture this level dynamically
l_attachments = []

l_PU_numbers = []
# Check if any arguments were passed
if len(sys.argv) > 1:
    # Print the argument (BATCH_DIR) passed from the batch file
    project_folder_path = sys.argv[1].strip('"')
    excel_project_template_name = "D131_PBŘ_XXX_přílohy.xlsx"
else:
    print("No arguments were passed to the script.")

# Get file paths from the user
excel_project_template_path = os.path.join(project_folder_path, excel_project_template_name)

workbook_data_only = openpyxl.load_workbook(excel_project_template_path, data_only=True)
sh_HL_PU_data_only = workbook_data_only["HL_PU"]

last_row_PU = 0

for row in range(4, sh_HL_PU_data_only.max_row+1):
    if sh_HL_PU_data_only.cell(row=row, column=1).value is not None:
        last_row_PU = row

for row in range(4, last_row_PU+1):
    if sh_HL_PU_data_only[f'H{row}'].value == "OB1":
        l_PU_numbers.append(sh_HL_PU_data_only[f'A{row}'].value)
    elif sh_HL_PU_data_only[f'H{row}'].value == "nevýrobní":
        l_PU_numbers.append(sh_HL_PU_data_only[f'A{row}'].value)
    elif sh_HL_PU_data_only[f'H{row}'].value == "nevýrobní (pv)":
        l_PU_numbers.append(sh_HL_PU_data_only[f'A{row}'].value)
    elif sh_HL_PU_data_only[f'H{row}'].value == "OB2":
        if "OB2" not in l_PU_numbers:
            l_PU_numbers.append("OB2")
    elif sh_HL_PU_data_only[f'H{row}'].value == "garáž I":
        l_PU_numbers.append(sh_HL_PU_data_only[f'A{row}'].value)
    elif sh_HL_PU_data_only[f'H{row}'].value == "garáž III":
        l_PU_numbers.append(sh_HL_PU_data_only[f'A{row}'].value)

# Define a function to perform on each paragraph within the section
def process_paragraph_in_section(paragraph_text):
    print(f"Processing paragraph: {paragraph_text}")
    if paragraph_text == "Stanovení kategorie staveb":
        l_attachments.append("SKT")
    elif paragraph_text == "Posouzení požární odolnosti stavebních konstrukcí":
        l_attachments.append("SK")
    elif paragraph_text == "Obsazení objektu osobami":
        l_attachments.append("OS")
    elif paragraph_text == "Posouzení nechráněných únikových cest":
        l_attachments.append("NÚC_nevýrobní")
    elif paragraph.text == "Posouzení částečně chráněných únikových cest":
        l_attachments.append("ČCHÚC_nevýrobní")
    elif paragraph_text == "Posouzení odstupových vzdáleností":
        l_attachments.append("PNP")
    elif paragraph_text == "Posouzení zařízení pro protipožární zásah":
        l_attachments.append("Hasiva")
    elif paragraph_text == "Posouzení změn užívání objektu":
        l_attachments.append("ZS")
    elif paragraph_text == "Stanovení požárního rizika PÚ":
        for el in l_PU_numbers:
            l_attachments.append(el)


# Iterate through paragraphs in the document
for paragraph in doc.paragraphs:
    # Check if the paragraph style is a heading
    if paragraph.style.name.startswith("Heading"):
        heading_text = paragraph.text.strip()

        # Capture the heading level (e.g., "Heading 1" = 1)
        heading_level = int(paragraph.style.name.split()[-1])

        # Check if this is the target heading
        if heading_text == target_heading:
            in_target_section = True
            target_level = heading_level
            continue  # Move to next paragraph after setting the section flag

        # If we're already in the target section and encounter a heading of the same or higher level, exit
        elif in_target_section and heading_level <= target_level:
            in_target_section = False

    # If in the target section, process the paragraph
    if in_target_section:
        process_paragraph_in_section(paragraph.text)

def excel_to_pdf_multiple_sheets(excel_path, l_attachments, output_folder, merged_pdf_name):
    pythoncom.CoInitialize()
    # Open Excel application
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False  # Make Excel run in the background

    # Open the existing Excel file
    workbook = excel.Workbooks.Open(excel_path)

    # List to store paths of individual PDF files
    pdf_files = []

    # Export each sheet in the l_attachments l_attachments to a separate PDF
    for sheet_name in l_attachments:
        try:
            # Get the specified sheet
            sheet = workbook.Sheets(sheet_name)

            # Create a unique PDF file path for each sheet
            pdf_file = f"{output_folder}\\{sheet_name}.pdf"
            pdf_files.append(pdf_file)

            # Export the specified sheet to PDF
            sheet.ExportAsFixedFormat(0, pdf_file)  # 0 = PDF format
            print(f"Exported {sheet_name} to {pdf_file}")

        except Exception as e:
            print(f"Failed to export sheet {sheet_name}: {e}")

    # Close the workbook and quit Excel
    workbook.Close(False)
    excel.Quit()

    # Explicitly release COM objects
    workbook = None
    excel = None
    pythoncom.CoUninitialize()

    # Merge the individual PDFs into a single PDF
    merger = PdfMerger()
    for pdf in pdf_files:
        merger.append(pdf)

    # Save the merged PDF file
    merged_pdf_path = os.path.join(output_folder, merged_pdf_name)
    merger.write(merged_pdf_path)
    merger.close()
    print(f"Merged PDF saved as {merged_pdf_path}")

    # Optionally delete individual PDFs after merging
    for pdf in pdf_files:
        os.remove(pdf)

excel_to_pdf_multiple_sheets(excel_path, l_attachments, pdf_output_folder, merged_pdf_name)
