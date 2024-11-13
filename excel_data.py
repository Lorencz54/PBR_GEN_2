import openpyxl
import os
import re
import sys
from docx import Document
import shutil
import subprocess
import time
import warnings

# Suppress only the specific openpyxl warning about Data Validation
warnings.filterwarnings("ignore", message=re.escape("Data Validation extension is not supported and will be removed"), category=UserWarning, module="openpyxl")



print("_________________________________________________connecting word and excel files")
# Check if any arguments were passed
if len(sys.argv) > 1:
    # Print the argument (BATCH_DIR) passed from the batch file
    project_folder_path = sys.argv[1].strip('"')
    excel_project_template_name = "D131_PBŘ_XXX_přílohy.xlsx"
else:
    print("No arguments were passed to the script.")

# Get file paths from the user
excel_project_template_path = os.path.join(project_folder_path, excel_project_template_name)
if "Lenovo" in excel_project_template_path:
    word_project_template_path = r"C:\Users\Lenovo\OneDrive\Projekty\Skamba\sablony\PBR\D131_PBŘ_nevýrobní_TZ.docx"
    cad_project_template_path = r"C:\Users\Lenovo\OneDrive\Projekty\Skamba\sablony\PBR\D132_PBŘ_XXX_výkresy.dwg"
else:
    word_project_template_path = r"C:\Users\kevin\OneDrive\Projekty\Skamba\sablony\PBR\D131_PBŘ_nevýrobní_TZ.docx"
    cad_project_template_path = r"C:\Users\kevin\OneDrive\Projekty\Skamba\sablony\PBR\D132_PBŘ_XXX_výkresy.dwg"
print("_________________________________________________files connected")
print("_________________________________________________renaming and loading excel file")
# Check if the excel_project_template_path file exists
if not os.path.exists(excel_project_template_path):
    # Set the new file path if the original template path does not exist
    path_parts = os.path.normpath(project_folder_path).split(os.sep)
    output_word_name = path_parts[-2] if len(path_parts) >= 2 else None

    # Remove the prefix pattern like "2024_53_" using regex
    if output_word_name:
        output_word_name = re.sub(r"^\d+_\d+_", "", output_word_name)

    if output_word_name != "sablony":
        output_excel_name_with_extension = f"D131_PBŘ_{output_word_name}_přílohy.xlsx" if output_word_name else None
        new_file_path = os.path.join(os.path.dirname(excel_project_template_path), output_excel_name_with_extension)
        excel_project_template_path = new_file_path
        print(f"Using new file path: {excel_project_template_path}")
else:
    # If the file exists, rename it as needed
    path_parts = os.path.normpath(project_folder_path).split(os.sep)
    output_word_name = path_parts[-2] if len(path_parts) >= 2 else None

    if output_word_name:
        output_word_name = re.sub(r"^\d+_\d+_", "", output_word_name)

    if output_word_name != "sablony":
        output_excel_name_with_extension = f"D131_PBŘ_{output_word_name}_přílohy.xlsx" if output_word_name else None
        new_file_path = os.path.join(os.path.dirname(excel_project_template_path), output_excel_name_with_extension)
        os.rename(excel_project_template_path, new_file_path)
        excel_project_template_path = new_file_path

# Proceed to load the workbook using the final excel_project_template_path
excel_workbook = openpyxl.load_workbook(excel_project_template_path)
sh_SKT = excel_workbook["SKT"]
workbook_data_only = openpyxl.load_workbook(excel_project_template_path, data_only=True)
sh_SKT_data_only = workbook_data_only["SKT"]
sh_HL_info_data_only = workbook_data_only["HL_info"]
sh_HL_PU_data_only = workbook_data_only["HL_PU"]
sh_HL_SK_data_only = workbook_data_only["HL_SK"]
sh_HL_ZS_data_only = workbook_data_only["HL_ZS"]

print("_________________________________________________excel file renamed and loaded")
print("_________________________________________________obtaining project information")
# údaje PD
nazev_projektu = sh_SKT['F4'].value
misto_projektu = sh_SKT['F5'].value
nazev_mista = sh_HL_info_data_only['A24'].value
kraj = sh_HL_info_data_only['D24'].value
ucel_stavby = sh_HL_info_data_only['G24'].value
charakter_stavby = sh_HL_info_data_only['K24'].value
#
# údaje zadavatele a zpracovatele PD
jmeno_zadavatel = sh_HL_info_data_only['C12'].value
adresa_zadavatel = sh_HL_info_data_only['C13'].value
jmeno_zpracovatel = sh_HL_info_data_only['C15'].value
adresa_zpracovatel = sh_HL_info_data_only['C16'].value
odpovedny_projektant = sh_HL_info_data_only['C17'].value
obor_autorizace = sh_HL_info_data_only['C18'].value
cislo_autorizace = sh_HL_info_data_only['C19'].value
zpracovatel_pd = sh_HL_info_data_only['C20'].value
tel_zpracovatel = sh_HL_info_data_only['C21'].value
mail_zpracovatel = sh_HL_info_data_only['C22'].value

if jmeno_zadavatel == odpovedny_projektant:
    spolecnost = " "
else:
    spolecnost = jmeno_zpracovatel
#
# základní požární informace
pozarni_vyska = "{:.2f}".format(sh_SKT['J14'].value).replace('.', ',')
pozarni_vyska_raw = sh_SKT['J14'].value
zastavena_plocha = "{:.2f}".format(sh_SKT['J13'].value).replace('.', ',')
k_system = sh_HL_info_data_only['K9'].value
pocet_NP_obj = sh_HL_info_data_only['D8'].value
pocet_PP_obj = sh_HL_info_data_only['D9'].value

pocet_osob = ""
pocet_osob_count = sh_SKT_data_only['L16'].value
if pocet_osob_count == 0:
    pocet_osob = "nula osob"
elif pocet_osob_count == 1:
    pocet_osob = "jedna osoba"
elif pocet_osob_count <= 4:
    pocet_osob = str(pocet_osob_count) + " osoby"
else:
    pocet_osob = str(pocet_osob_count) + " osob"

kategorie = 0
kategorie_count = sh_HL_info_data_only['K7'].value
if kategorie_count == "K I":
    kategorie = "I"
elif kategorie_count == "K II":
    kategorie = "II"
elif kategorie_count == "K III":
    kategorie = "III"
else:
    pass

trida_vyuziti = "bez třídy"
trida_vyuziti_count = sh_HL_info_data_only['K8'].value

if trida_vyuziti_count == "T1":
    trida_vyuziti = "první třída"
elif trida_vyuziti_count == "T2":
    trida_vyuziti = "druhá třída"
elif trida_vyuziti_count == "T3":
    trida_vyuziti = "třetí třída"
elif trida_vyuziti_count == "T4":
    trida_vyuziti = "čtvrtá třída"
elif trida_vyuziti_count == "T5":
    trida_vyuziti = "pátá třída"
else:
    pass
#
# základní popis objektu
predmet_PBR = sh_HL_info_data_only['O3'].value
umisteni_obj = sh_HL_info_data_only['O10'].value
zakladni_popis_obj = sh_HL_info_data_only['O16'].value
#
print("_________________________________________________project information obtained")
print("_________________________________________________generating and filling CSN l_attachments")
# normy ČSN
l_chosen_CSN = []
for row in range(26, 32):  # Adjust the range as needed
    raw_value = sh_HL_info_data_only[f'D{row}'].value  # Adjust 'A' to your column
    if str(raw_value) != "None":  # True
        l_chosen_CSN.append(raw_value)  # Use the position (row number) as the value

for row in range(26, 32):  # Adjust the range as needed
    raw_value = sh_HL_info_data_only[f'H{row}'].value  # Adjust 'A' to your column
    if str(raw_value) != "None":  # True
        l_chosen_CSN.append(raw_value)  # Use the position (row number) as the value
#
print("_________________________________________________all CSN listed")
print("_________________________________________________obtaining object information")
# informace o zmene stavby
zmena_stavby = sh_HL_info_data_only['C45'].value
skupina_zmeny_stavby = sh_HL_info_data_only['H45'].value
popis_navrzenych_zmen = sh_HL_ZS_data_only['A3'].value
#
# informace o OB objektu
objekt_pro_bydleni = sh_HL_info_data_only['C34'].value
pocet_OB = sh_HL_info_data_only['H34'].value
podkrovni_NP = sh_HL_info_data_only['M34'].value
rekreacni_objekt = sh_HL_info_data_only['H35'].value
#
# informace o sousedních objektech
sousedni_objekt = sh_HL_info_data_only['C36'].value
#
# informace o garážích
typ_garaze = sh_HL_info_data_only['C37'].value
pristresek = sh_HL_info_data_only['H37'].value
garaz_soucasti_RD = sh_HL_info_data_only['M37'].value
pocet_stani = sh_HL_info_data_only['H38'].value
vice_nez_50_procent_sten = sh_HL_info_data_only['M38'].value
konstrukce_garaze = sh_HL_info_data_only['H39'].value
druh_paliv = sh_HL_info_data_only['M39'].value
skupina_garaze = sh_HL_info_data_only['M40'].value

instalace_FVE = sh_HL_info_data_only['C42'].value
umisteni_FVE = sh_HL_info_data_only['H42'].value
FVE_baterie = sh_HL_info_data_only['M42'].value
celkovy_vykon_FVE = sh_HL_info_data_only['H43'].value
vyvin_tepla_FVE = sh_HL_info_data_only['M43'].value
print("_________________________________________________object information obtained")
print("_________________________________________________generating lists for constructions")
# konstrukce
l_svisle_nosne = []
l_vodorovne_nosne = []
l_svisle_nenosne = []
l_vodorovne_nenosne = []
l_konstrukce_strechy = []
l_stresni_krytiny = []
l_tepelne_izolace = []
l_schodiste = []
l_podlahy = []
l_vyplne_otvoru = []
l_vnejsi_povrchy = []

l_material_svisle_nosne = []
l_material_vodorovne_nosne = []
l_material_svisle_nenosne = []
l_material_vodorovne_nenosne = []
l_material_konstrukce_strechy = []
l_material_stresni_krytiny = []
l_material_tepelne_izolace = []
l_material_schodiste = []
l_material_podlahy = []
l_material_vyplne_otvoru = []
l_material_vnejsi_povrchy = []

l_tridy_svisle_nosne = []
l_tridy_vodorovne_nosne = []
l_tridy_svisle_nenosne = []
l_tridy_vodorovne_nenosne = []
l_tridy_konstrukce_strechy = []
l_tridy_stresni_krytiny = []
l_tridy_tepelne_izolace = []
l_tridy_schodiste = []
l_tridy_podlahy = []
l_tridy_vyplne_otvoru = []
l_tridy_vnejsi_povrchy = []

l_odolnost_svisle_nosne = []
l_odolnost_vodorovne_nosne = []
l_odolnost_svisle_nenosne = []
l_odolnost_vodorovne_nenosne = []
l_odolnost_konstrukce_strechy = []
l_odolnost_schodiste = []

l_DP_svisle_nosne = []
l_DP_vodorovne_nosne = []
l_DP_svisle_nenosne = []
l_DP_vodorovne_nenosne = []
l_DP_konstrukce_strechy = []

d_DP_konstrukce = [l_DP_svisle_nosne, l_DP_vodorovne_nosne, l_DP_svisle_nenosne, l_DP_vodorovne_nenosne, l_DP_konstrukce_strechy]

l_pozarne_delici_svisle_nosne = []
l_pozarne_delici_vodorovne_nosne = []
l_pozarne_delici_svisle_nenosne = []
l_pozarne_delici_vodorovne_nenosne = []
l_pozarne_delici_konstrukce_strechy = []

d_pozarne_delici_konstrukce = [l_pozarne_delici_svisle_nosne, l_pozarne_delici_vodorovne_nosne, l_pozarne_delici_svisle_nenosne, l_pozarne_delici_vodorovne_nenosne, l_pozarne_delici_konstrukce_strechy]
#
# listy pro změny staveb
l_clanky_stavebnich_uprav_I = []
l_popis_stavebnich_uprav = []
#
last_row_zmeny_staveb = 0
for row in range(6, sh_HL_ZS_data_only.max_row+1):
    if sh_HL_ZS_data_only.cell(row=row, column=1).value is not None:
        last_row_zmeny_staveb = row

for row in range(6, last_row_zmeny_staveb+1):
    l_clanky_stavebnich_uprav_I.append(sh_HL_ZS_data_only[f'A{row}'].value)
    l_popis_stavebnich_uprav.append(sh_HL_ZS_data_only[f'C{row}'].value)

last_row_konstrukce = 0
for row in range(5, sh_HL_SK_data_only.max_row+1):
    if sh_HL_SK_data_only.cell(row=row, column=1).value is not None:
        last_row_konstrukce = row

for row in range(5, last_row_konstrukce+1):
    if sh_HL_SK_data_only[f'F{row}'].value == "svislá nosná":
        if sh_HL_SK_data_only[f'I{row}'] not in l_material_svisle_nosne:
            l_material_svisle_nosne.append(sh_HL_SK_data_only[f'I{row}'])
            l_tridy_svisle_nosne.append(sh_HL_SK_data_only[f'K{row}'])
            l_odolnost_svisle_nosne.append(f"{sh_HL_SK_data_only[f'L{row}'].value} {sh_HL_SK_data_only[f'M{row}'].value} {sh_HL_SK_data_only[f'N{row}'].value}")
            l_DP_svisle_nosne.append(sh_HL_SK_data_only[f'N{row}'].value)
            l_pozarne_delici_svisle_nosne.append(sh_HL_SK_data_only[f'O{row}'].value)
        l_svisle_nosne.append(sh_HL_SK_data_only[f'A{row}'])
    elif sh_HL_SK_data_only[f'F{row}'].value == "vodorovná nosná":
        if sh_HL_SK_data_only[f'I{row}'] not in l_material_vodorovne_nosne:
            l_material_vodorovne_nosne.append(sh_HL_SK_data_only[f'I{row}'])
            l_tridy_vodorovne_nosne.append(sh_HL_SK_data_only[f'K{row}'])
            l_odolnost_vodorovne_nosne.append(f"{sh_HL_SK_data_only[f'L{row}'].value} {sh_HL_SK_data_only[f'M{row}'].value} {sh_HL_SK_data_only[f'N{row}'].value}")
            l_DP_vodorovne_nosne.append(sh_HL_SK_data_only[f'N{row}'].value)
            l_pozarne_delici_vodorovne_nosne.append(sh_HL_SK_data_only[f'O{row}'].value)
        l_vodorovne_nosne.append(sh_HL_SK_data_only[f'A{row}'])
    elif sh_HL_SK_data_only[f'F{row}'].value == "svislá nenosná":
        if sh_HL_SK_data_only[f'I{row}'] not in l_material_svisle_nenosne:
            l_material_svisle_nenosne.append(sh_HL_SK_data_only[f'I{row}'])
            l_tridy_svisle_nenosne.append(sh_HL_SK_data_only[f'K{row}'])
            l_odolnost_svisle_nenosne.append(f"{sh_HL_SK_data_only[f'L{row}'].value} {sh_HL_SK_data_only[f'M{row}'].value} {sh_HL_SK_data_only[f'N{row}'].value}")
            l_DP_svisle_nenosne.append(sh_HL_SK_data_only[f'N{row}'].value)
            l_pozarne_delici_svisle_nenosne.append(sh_HL_SK_data_only[f'O{row}'].value)
        l_svisle_nenosne.append(sh_HL_SK_data_only[f'A{row}'])
    elif sh_HL_SK_data_only[f'F{row}'].value == "vodorovná nenosná":
        if sh_HL_SK_data_only[f'I{row}'] not in l_material_vodorovne_nenosne:
            l_material_vodorovne_nenosne.append(sh_HL_SK_data_only[f'I{row}'])
            l_tridy_vodorovne_nenosne.append(sh_HL_SK_data_only[f'K{row}'])
            l_odolnost_vodorovne_nenosne.append(f"{sh_HL_SK_data_only[f'L{row}'].value} {sh_HL_SK_data_only[f'M{row}'].value} {sh_HL_SK_data_only[f'N{row}'].value}")
            l_DP_vodorovne_nenosne.append(sh_HL_SK_data_only[f'N{row}'].value)
            l_pozarne_delici_vodorovne_nenosne.append(sh_HL_SK_data_only[f'O{row}'].value)
        l_vodorovne_nenosne.append(sh_HL_SK_data_only[f'A{row}'])
    elif sh_HL_SK_data_only[f'F{row}'].value == "konstrukce střechy":
        if sh_HL_SK_data_only[f'I{row}'] not in l_material_konstrukce_strechy:
            l_material_konstrukce_strechy.append(sh_HL_SK_data_only[f'I{row}'])
            l_tridy_konstrukce_strechy.append(sh_HL_SK_data_only[f'K{row}'])
            l_odolnost_konstrukce_strechy.append(f"{sh_HL_SK_data_only[f'L{row}'].value} {sh_HL_SK_data_only[f'M{row}'].value} {sh_HL_SK_data_only[f'N{row}'].value}")
            l_DP_konstrukce_strechy.append(sh_HL_SK_data_only[f'N{row}'].value)
            l_pozarne_delici_konstrukce_strechy.append(sh_HL_SK_data_only[f'O{row}'].value)
        l_konstrukce_strechy.append(sh_HL_SK_data_only[f'A{row}'])
    elif sh_HL_SK_data_only[f'F{row}'].value == "střešní krytina":
        if sh_HL_SK_data_only[f'I{row}'] not in l_material_stresni_krytiny:
            l_material_stresni_krytiny.append(sh_HL_SK_data_only[f'I{row}'])
            l_tridy_stresni_krytiny.append(sh_HL_SK_data_only[f'K{row}'])
        l_stresni_krytiny.append(sh_HL_SK_data_only[f'A{row}'])
    elif sh_HL_SK_data_only[f'F{row}'].value == "tepelná izolace":
        if sh_HL_SK_data_only[f'I{row}'] not in l_material_tepelne_izolace:
            l_material_tepelne_izolace.append(sh_HL_SK_data_only[f'I{row}'])
            l_tridy_tepelne_izolace.append(sh_HL_SK_data_only[f'K{row}'])
        l_tepelne_izolace.append(sh_HL_SK_data_only[f'A{row}'])
    elif sh_HL_SK_data_only[f'F{row}'].value == "schodiště":
        if sh_HL_SK_data_only[f'I{row}'] not in l_material_schodiste:
            l_material_schodiste.append(sh_HL_SK_data_only[f'I{row}'])
            l_tridy_schodiste.append(sh_HL_SK_data_only[f'K{row}'])
            l_odolnost_schodiste.append(f"{sh_HL_SK_data_only[f'L{row}'].value} {sh_HL_SK_data_only[f'M{row}'].value} {sh_HL_SK_data_only[f'N{row}'].value}")
        l_schodiste.append(sh_HL_SK_data_only[f'A{row}'])
    elif sh_HL_SK_data_only[f'F{row}'].value == "podlaha":
        if sh_HL_SK_data_only[f'I{row}'] not in l_material_podlahy:
            l_material_podlahy.append(sh_HL_SK_data_only[f'I{row}'])
            l_tridy_podlahy.append(sh_HL_SK_data_only[f'K{row}'])
        l_podlahy.append(sh_HL_SK_data_only[f'A{row}'])
    elif sh_HL_SK_data_only[f'F{row}'].value == "výplň otvoru":
        if sh_HL_SK_data_only[f'I{row}'] not in l_material_vyplne_otvoru:
            l_material_vyplne_otvoru.append(sh_HL_SK_data_only[f'I{row}'])
            l_tridy_vyplne_otvoru.append(sh_HL_SK_data_only[f'K{row}'])
        l_vyplne_otvoru.append(sh_HL_SK_data_only[f'A{row}'])
    elif sh_HL_SK_data_only[f'F{row}'].value == "vnější povrch":
        if sh_HL_SK_data_only[f'I{row}'] not in l_material_vnejsi_povrchy:
            l_material_vnejsi_povrchy.append(sh_HL_SK_data_only[f'I{row}'])
            l_tridy_vnejsi_povrchy.append(sh_HL_SK_data_only[f'K{row}'])
        l_vnejsi_povrchy.append(sh_HL_SK_data_only[f'A{row}'])
print("_________________________________________________all constructions listed")
print("_________________________________________________generating and filling PUs l_attachments")
d_PU_types = []

last_row_PU = 0
for row in range(4, sh_HL_PU_data_only.max_row+1):
    if sh_HL_PU_data_only.cell(row=row, column=1).value is not None:
        last_row_PU = row
for row in range(4, last_row_PU+1):
    l_PU_properties = []
    if sh_HL_PU_data_only[f'H{row}'].value == "OB1":
        l_PU_properties.append("OB1") # přidá typ PÚ [0]
        l_PU_properties.append(sh_HL_PU_data_only[f'A{row}'].value.replace("_", " ")) # přidá označení PÚ [1]
        l_PU_properties.append(sh_HL_PU_data_only[f'C{row}'].value) # přidá název PÚ [2]
        l_PU_properties.append("{:.2f}".format(sh_HL_PU_data_only[f'L{row}'].value).replace('.', ',')) # přidá pv PÚ [3]
        l_PU_properties.append(sh_HL_PU_data_only[f'N{row}'].value)  # přidá SPB PÚ [4]
        sh_PU = workbook_data_only[sh_HL_PU_data_only[f'A{row}'].value] # definuje excelový l_attachments PÚ
        l_PU_properties.append("{:.2f}".format(sh_PU["T5"].value).replace('.', ','))  # přidá ps PÚ [5]
        d_PU_types.append(l_PU_properties)

    elif sh_HL_PU_data_only[f'H{row}'].value == "nevýrobní":
        l_PU_properties.append("nevýrobní") # přidá typ PÚ [0]
        l_PU_properties.append(sh_HL_PU_data_only[f'A{row}'].value.replace("_", " "))  # přidá označení PÚ [1]
        l_PU_properties.append(sh_HL_PU_data_only[f'C{row}'].value) # přidá název PÚ [2]
        l_PU_properties.append("{:.2f}".format(sh_HL_PU_data_only[f'L{row}'].value).replace('.', ',')) # přidá pv´ PÚ [3]
        l_PU_properties.append(sh_HL_PU_data_only[f'N{row}'].value)  # přidá SPB PÚ [4]
        sh_PU = workbook_data_only[sh_HL_PU_data_only[f'A{row}'].value] # definuje excelový l_attachments PÚ
        l_PU_properties.append("{:.2f}".format(sh_PU["M6"].value).replace('.', ','))  # přidá ps PÚ [5]
        l_PU_properties.append("{:.2f}".format(sh_PU["M5"].value).replace('.', ','))  # přidá pn PÚ [6]
        l_PU_properties.append("{:.2f}".format(sh_HL_PU_data_only[f'M{row}'].value).replace('.', ','))  # přidá součinitel a PÚ [7]
        d_PU_types.append(l_PU_properties)

    elif sh_HL_PU_data_only[f'H{row}'].value == "nevýrobní (pv)":
        l_PU_properties.append("nevýrobní (pv)") # přidá typ PÚ [0]
        l_PU_properties.append(sh_HL_PU_data_only[f'A{row}'].value.replace("_", " ")) # přidá označení PÚ [1]
        l_PU_properties.append(sh_HL_PU_data_only[f'C{row}'].value) # přidá název PÚ [2]
        l_PU_properties.append("{:.2f}".format(sh_HL_PU_data_only[f'L{row}'].value).replace('.', ',')) # přidá pv´ PÚ [3]
        l_PU_properties.append(sh_HL_PU_data_only[f'N{row}'].value)  # přidá SPB PÚ [4]
        sh_PU = workbook_data_only[sh_HL_PU_data_only[f'A{row}'].value] # definuje excelový l_attachments PÚ
        l_PU_properties.append("{:.2f}".format(sh_PU["T7"].value).replace('.', ','))  # přidá ps PÚ [5]
        l_PU_properties.append("{:.2f}".format(sh_PU["C2"].value).replace('.', ','))  # přidá pv PÚ [6]
        l_PU_properties.append("{:.2f}".format(sh_PU["C3"].value).replace('.', ','))  # přidá a PÚ [7]
        d_PU_types.append(l_PU_properties)

    elif sh_HL_PU_data_only[f'H{row}'].value == "OB2":
        l_PU_properties.append("OB2") # přidá typ PÚ [0]
        l_PU_properties.append(sh_HL_PU_data_only[f'A{row}'].value.replace("_", " ")) # přidá označení PÚ [1]
        l_PU_properties.append(sh_HL_PU_data_only[f'C{row}'].value) # přidá název PÚ [2]
        l_PU_properties.append("{:.2f}".format(sh_HL_PU_data_only[f'L{row}'].value).replace('.', ',')) # přidá pv PÚ [3]
        l_PU_properties.append(sh_HL_PU_data_only[f'N{row}'].value)  # přidá SPB PÚ [4]
        d_PU_types.append(l_PU_properties)

    elif sh_HL_PU_data_only[f'H{row}'].value == "garáž I":
        l_PU_properties.append("garáž I") # přidá typ PÚ [0]
        l_PU_properties.append(sh_HL_PU_data_only[f'A{row}'].value.replace("_", " ")) # přidá označení PÚ [1]
        l_PU_properties.append(sh_HL_PU_data_only[f'C{row}'].value) # přidá název PÚ [2]
        l_PU_properties.append("{:.2f}".format(sh_HL_PU_data_only[f'L{row}'].value).replace('.', ',')) # přidá pv´ PÚ [3]
        l_PU_properties.append(sh_HL_PU_data_only[f'N{row}'].value)  # přidá SPB PÚ [4]
        sh_PU = workbook_data_only[sh_HL_PU_data_only[f'A{row}'].value] # definuje excelový l_attachments PÚ
        l_PU_properties.append("{:.2f}".format(sh_PU["T7"].value).replace('.', ','))  # přidá ps PÚ [5]
        l_PU_properties.append("{:.2f}".format(sh_PU["C2"].value).replace('.', ','))  # přidá pv PÚ [6]
        l_PU_properties.append("{:.2f}".format(sh_PU["C3"].value).replace('.', ','))  # přidá a PÚ [7]
        d_PU_types.append(l_PU_properties)

    elif sh_HL_PU_data_only[f'H{row}'].value == "garáž III":
        l_PU_properties.append("garáž III")  # přidá typ PÚ [0]
        l_PU_properties.append(sh_HL_PU_data_only[f'A{row}'].value.replace("_", " "))  # přidá označení PÚ [1]
        l_PU_properties.append(sh_HL_PU_data_only[f'C{row}'].value)  # přidá název PÚ [2]
        l_PU_properties.append("{:.2f}".format(sh_HL_PU_data_only[f'L{row}'].value).replace('.', ','))  # přidá pv´ PÚ [3]
        l_PU_properties.append(sh_HL_PU_data_only[f'N{row}'].value)  # přidá SPB PÚ [4]
        sh_PU = workbook_data_only[sh_HL_PU_data_only[f'A{row}'].value]  # definuje excelový l_attachments PÚ
        l_PU_properties.append("{:.2f}".format(sh_PU["T13"].value).replace('.', ','))  # přidá ps PÚ [5]
        l_PU_properties.append("{:.2f}".format(sh_PU["B2"].value).replace('.', ','))  # přidá pv PÚ [6]
        l_PU_properties.append("{:.2f}".format(sh_PU["B3"].value).replace('.', ','))  # přidá a PÚ [8]
        l_PU_properties.append("{:.2f}".format(sh_PU["M8"].value).replace('.', ','))  # přidá skutečnou kapacitu vozů garáže [8]
        l_PU_properties.append("{:.2f}".format(sh_PU["R8"].value).replace('.', ',')) # přidá mezní počet stání hromadné garáže [9]
        d_PU_types.append(l_PU_properties)
print("_________________________________________________all PUs listed")
print("_________________________________________________creating and renaming output document")

def terminate_process(process_name):
    try:
        # Use taskkill command to terminate the process
        subprocess.call(['taskkill', '/F', '/IM', process_name])
        time.sleep(1)  # Give it a moment to close
    except Exception as e:
        pass  # Handle silently or log if needed

# Specify the name of the process you want to kill (e.g., 'WINWORD.EXE' for Word)
process_name = 'WINWORD.EXE'  # Change this to your process name

# Assume word_project_template_path and excel_project_template_path are defined and valid
word_document = Document(word_project_template_path)


if output_word_name == "sablony":
    output_word_name_with_extension = f"D131_PBŘ_test_TZ.docx" if output_word_name else None
    output_word_path = os.path.join(project_folder_path, output_word_name_with_extension)

else:
    output_word_name_with_extension = f"D131_PBŘ_{output_word_name}_TZ.docx" if output_word_name else None
    output_cad_name_with_extension = f"D132_PBŘ_{output_word_name}_výkresy.dwg" if output_word_name else None
    # Output the result
    output_word_path = os.path.join(project_folder_path, output_word_name_with_extension)

    # Construct the destination file path
    output_cad_path = os.path.join(project_folder_path, output_cad_name_with_extension)

    # Check if the source and destination are the same or if the file already exists in the destination folder
    if os.path.abspath(cad_project_template_path) == os.path.abspath(output_cad_path):
        print("Source and destination paths are the same. File copy skipped.")
    elif os.path.exists(output_cad_path):
        print("File already exists in the destination folder.")
        # Optional: Rename the file to avoid overwriting
        new_destination_file = os.path.join(project_folder_path, f"copy_of_{os.path.basename(cad_project_template_path)}")
        shutil.copy(cad_project_template_path, new_destination_file)
        print(f"File copied as: {new_destination_file}")
    else:
        shutil.copy(cad_project_template_path, output_cad_path)
        print("File copied successfully!")

try:
    word_document.save(output_word_path)
except Exception:
    terminate_process(process_name)
    word_document.save(output_word_path)  # Try saving again after terminating the process
print("_________________________________________________Output document created and renamed")
