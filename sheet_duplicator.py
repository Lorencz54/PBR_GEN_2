import xlwings as xw
import openpyxl
import sys
import os

if len(sys.argv) > 1:
    # Print the argument (BATCH_DIR) passed from the batch file
    batch_dir = sys.argv[1]
    print(f"Batch directory received: {batch_dir}")
    batch_dir = sys.argv[1].strip('"')
    excel_filename = "D131_PBŘ_XXX_přílohy.xlsx"
else:
    print("No arguments were passed to the script.")

excel_file = os.path.join(batch_dir, excel_filename)
workbook = openpyxl.load_workbook(excel_file)
sh_SKT = workbook["SKT"]
workbook_data_only = openpyxl.load_workbook(excel_file, data_only=True)
sh_SKT_data_only = workbook_data_only["SKT"]
sh_HL_info_data_only = workbook_data_only["HL_info"]
sh_HL_PU_data_only = workbook_data_only["HL_PU"]
sh_HL_SK_data_only = workbook_data_only["HL_SK"]
d_PU_types = []

last_row_PU = 0
for row in range(4, sh_HL_PU_data_only.max_row+1):
    if sh_HL_PU_data_only.cell(row=row, column=1).value is not None:
        last_row_PU = row
for row in range(4, last_row_PU+1):
    l_PU_properties = []
    if sh_HL_PU_data_only[f'H{row}'].value == "OB1":
        l_PU_properties.append("OB1") # přidá typ PÚ [0]
        l_PU_properties.append(sh_HL_PU_data_only[f'A{row}'].value.replace("_", " ")) # přidá označení PÚ [1]
        d_PU_types.append(l_PU_properties)

    elif sh_HL_PU_data_only[f'H{row}'].value == "nevýrobní":
        l_PU_properties.append("nevýrobní") # přidá typ PÚ [0]
        l_PU_properties.append(sh_HL_PU_data_only[f'A{row}'].value.replace("_", " "))  # přidá označení PÚ [1]
        d_PU_types.append(l_PU_properties)

    elif sh_HL_PU_data_only[f'H{row}'].value == "nevýrobní (pv)":
        l_PU_properties.append("nevýrobní (pv)") # přidá typ PÚ [0]
        l_PU_properties.append(sh_HL_PU_data_only[f'A{row}'].value.replace("_", " ")) # přidá označení PÚ [1]
        d_PU_types.append(l_PU_properties)

    elif sh_HL_PU_data_only[f'H{row}'].value == "OB2":
        l_PU_properties.append("OB2") # přidá typ PÚ [0]
        l_PU_properties.append(sh_HL_PU_data_only[f'A{row}'].value.replace("_", " ")) # přidá označení PÚ [1]
        d_PU_types.append(l_PU_properties)

    elif sh_HL_PU_data_only[f'H{row}'].value == "garáž I":
        l_PU_properties.append("garáž I") # přidá typ PÚ [0]
        l_PU_properties.append(sh_HL_PU_data_only[f'A{row}'].value.replace("_", " ")) # přidá označení PÚ [1]
        d_PU_types.append(l_PU_properties)

    elif sh_HL_PU_data_only[f'H{row}'].value == "garáž III":
        l_PU_properties.append("garáž III")  # přidá typ PÚ [0]
        l_PU_properties.append(sh_HL_PU_data_only[f'A{row}'].value.replace("_", " "))  # přidá označení PÚ [1]
        d_PU_types.append(l_PU_properties)

    elif sh_HL_PU_data_only[f'H{row}'].value == "instalační šachta":
        l_PU_properties.append("instalační šachta")  # přidá typ PÚ [0]
        l_PU_properties.append(sh_HL_PU_data_only[f'A{row}'].value.replace("_", " "))  # přidá označení PÚ [1]
        d_PU_types.append(l_PU_properties)

# Variable to track if the "instalační šachty" sheet has been duplicated
duplicated_instalacni_sachty = False
duplicated_OB2 = False

# Open the excel_workbook once
with xw.App(visible=False) as app:  # Set visible=True to see what's happening
    workbook = app.books.open(excel_file)
    # Check conditions for duplicating sheets
    for list in d_PU_types:
        if list[0] == "instalační šachta" and duplicated_instalacni_sachty is False:
            duplicated_instalacni_sachty = True  # Mark as duplicated
            # Copy the sheet and rename it
            sheet_to_copy = workbook.sheets[list[0]]
            sheet_to_copy.api.Copy(Before=workbook.sheets[0].api)

            # Find and rename the newly copied sheet
            for sheet in workbook.sheets:
                if sheet.name == sheet_to_copy.name + " (2)":  # Default copied sheet name
                    sheet.name = list[1]
                    break

        elif list[0] == "OB2" and duplicated_OB2 is False:
            # Copy the sheet and rename it
            sheet_to_copy = workbook.sheets[list[0]]
            sheet_to_copy.api.Copy(Before=workbook.sheets[0].api)

            # Find and rename the newly copied sheet
            for sheet in workbook.sheets:
                if sheet.name == sheet_to_copy.name + " (2)":  # Default copied sheet name
                    sheet.name = list[1]
                    break
        elif list[0] != "instalační šachta" or list[0] != "OB2":
            # Copy the sheet and rename it
            sheet_to_copy = workbook.sheets[list[0]]
            sheet_to_copy.api.Copy(Before=workbook.sheets[0].api)

            # Find and rename the newly copied sheet
            for sheet in workbook.sheets:
                if sheet.name == sheet_to_copy.name + " (2)":  # Default copied sheet name
                    sheet.name = list[1]
                    break
    workbook.save()
    workbook.close()
